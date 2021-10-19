from selenium import webdriver
from selenium.webdriver.common.by import By
from urllib import request as rq
from selenium.webdriver.common.keys import Keys
import saveToExcel as ex
import openpyxl as xl

url = 'https://finance.naver.com/sise/lastsearch2.naver'
browser = webdriver.Chrome()
browser.get(url)

#
# no = browser.find_elements(By.CLASS_NAME, 'no')
# tltle = browser.find_elements(By.CLASS_NAME, 'tltle')
# numbers = browser.find_elements(By.CLASS_NAME, 'number')




numberSelector = '#contentarea > div.box_type_l > table > tbody > tr > td'
todaysStock = browser.find_elements(By.CSS_SELECTOR, numberSelector)

for idx, stock in enumerate(todaysStock):
    print(idx, stock.text)

    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = '테스트'

    col_names = ['순위', '종목명', '검색비율', '현재가', '전일비', '등락률',
                 '거래량', '시가', '고가', '저가', 'PER', 'ROE']

    for seq, name in enumerate(col_names):
        sheet.cell(row=1, column=seq+1, value=name)

    for idx, mode in enumerate(todaysStock):
        todaysStock = stock.text

    row_no = 2

    for i in range():
        sheet.cell(row=row_no + 12, column=seq + 1, value='')
          #
          # if todaysStock == "":
          #   continue
        wb.save("C:/chh_scraping/download/naverStock.xlsx")
        wb.close()




        # weatherData = [
        #     (mode.string, tmEfs[idx].string, wfs[idx].string, tmns[idx].string, tmxs[idx].string, rnSts[idx].string)]
        #
        # row_no = 2
        # for n, rows in enumerate(weatherData):
        #     for seq, value in enumerate(rows):
        #         sheet.cell(row=row_no + n, column=seq + 1, value=value)
        #         sheet.append([mode.string, tmEfs[idx].string, wfs[idx].string, tmns[idx].string, tmxs[idx].string,
        #                       rnSts[idx].string])

browser.close()

















