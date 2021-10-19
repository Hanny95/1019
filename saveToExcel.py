from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

def write2Excel(dataList):

    dir = 'C:/chh_scraping/download/'
    fileName = 'todaysStock.xlsx'

    # 파일 여부 확인
    isFile = os.path.isfile(dir + fileName)

    if isFile:
       wb = load_workbook(dir + fileName)

       ws = wb['new_sheet']
       ws[ws.max_row]   # 가장 끝에 있는 열
       ws.append(dataList)


       wb.save(dir + fileName)
       wb.close()

    else:
        wb = Workbook(dir + fileName)

        ws = wb.create_sheet('new_sheet')  # workSheet 생성
        ws.append(dataList)  # 리스트

        wb.save(dir + fileName)   # Save (저장)
        wb.close()                # 자원해제